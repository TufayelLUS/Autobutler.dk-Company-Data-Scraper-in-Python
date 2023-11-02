import requests
import re
import json
from bs4 import BeautifulSoup as bs
import openpyxl
import os
import datetime

input_file = "postcode.xlsx"
record_file_name = "leads"
postcode_list = []
old_urls = []
s = requests.Session()


def loadPastRecords():
    global old_urls
    if os.path.exists(record_file_name + ".xlsx"):
        wb = openpyxl.load_workbook(record_file_name + ".xlsx")
        sht = wb.active
        for row in sht.iter_rows(min_row=2, values_only=True):
            url = row[3]
            old_urls.append(url)


def saveData(dataset, output_file):
    output_file = "{}.xlsx".format(output_file)  # this will be xlsx file name
    fieldnames = ["Date Recorded", "Postcode", "Company Name",
                  "Company URL"]  # change your header list here
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        sht = wb.active
        sht.append(dataset)
        wb.save(output_file)
    else:
        wb = openpyxl.Workbook()
        sht = wb.active
        sht.append(fieldnames)
        sht.append(dataset)
        wb.save(output_file)


def listResults(postcode):
    print("Getting data for postcode {}".format(postcode))
    link = "https://www.autobutler.dk/autovaerksted/{}?per_page=all".format(postcode)
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36'
    }
    resp = requests.get(link, headers=headers).content
    soup = bs(resp, 'html.parser')
    results = soup.findAll('div', {'class': 'mechanic-card'})
    for result in results:
        company_name = result.find(
            'div', {'class': 'mechanic-name'}).a.text.strip()
        company_url = "https://www.autobutler.dk" + \
            result.find('div', {'class': 'mechanic-name'}).a.get('href')
        print("PostCode: {}".format(postcode))
        print("Company Name: {}".format(company_name))
        print("Company URL: {}".format(company_url))
        if checkDuplicateLink(company_url):
            print("Skipped!")
            continue
        dataset = [datetime.datetime.now().strftime(
            "%m/%d/%Y"), postcode, company_name, company_url]
        saveData(dataset, record_file_name)
        print("Saved in excel sheet ...")


def LoadPostCodes():
    global postcode_list
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cell_data = row[0]
        if cell_data is not None:
            postcode_list.append(str(cell_data))


def checkDuplicateLink(link):
    if link.strip() == "":
        return False
    return link in old_urls


if __name__ == "__main__":
    loadPastRecords()
    LoadPostCodes()
    for postcode in postcode_list:
        if postcode == "":
            continue
        listResults(postcode)
